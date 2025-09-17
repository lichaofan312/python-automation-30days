from openpyxl import load_workbook, Workbook


def excel_auto_process(input_path, output_path):
    print('excel_auto_process')
    # 1 加载待处理的excel 文件
    workbook = load_workbook(input_path)
    # 选择活跃表格
    sheet = workbook.active
    # 创建一个新的excel, 用于保存结果
    new_wb = Workbook()
    new_sheet = new_wb.active
    new_sheet.append(['姓名', '部门', '销售额', '提成'])
    # 遍历原表格数据(跳过表头)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(row)
        name, dept, sale = row[0], row[1], row[2]
        if sale > 30000:
            commision = sale * 0.05  # 提成
            new_sheet.append([name, dept, sale, round(commision, 2)])

    # 保存文件
    new_wb.save(output_path)
    print(f'Excel处理完成， 结果已保存到: {output_path}')

print(f'__name__:{__name__}')
if __name__ == '__main__':
    input_excel = "员工销售数据.xlsx"
    output_excel = "高销售额员工报表.xlsx"
    excel_auto_process(input_excel, output_excel)
