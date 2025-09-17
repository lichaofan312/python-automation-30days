# 从 Excel 中读取 “客户邮箱” 和 “对应的 PDF 报表名”，自动发送带附件的邮件，邮件内容可自定义。
import os.path
import smtplib
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from openpyxl import load_workbook


def send_auto_email(smtp_server, smtp_port, sender_email, sender_password, root_path, excel_path):
    # 1. 加载Excel中的客户信息（假设列：客户姓名、客户邮箱、PDF附件路径）
    excel_entire_path = os.path.join(root_path, excel_path)
    print(excel_entire_path)
    wb = load_workbook(excel_entire_path)
    ws = wb.active

    # 2. 连接邮件服务器（以QQ邮箱为例，其他邮箱需改服务器和端口）
    try:
        server = smtplib.SMTP_SSL(smtp_server, smtp_port)  # SSL加密连接
        server.login(sender_email, sender_password)  # 登录邮箱（密码用“授权码”，不是登录密码！）
        print("邮箱登录成功！")
    except Exception as e:
        print(f"邮箱登录失败：{str(e)}")
        return

    # 3. 遍历客户信息，逐个发送邮件
    for row in ws.iter_rows(min_row=2, values_only=True):
        client_name, client_email, pdf_path = row[0], row[1], row[2]
        if not (client_name and client_email and pdf_path):
            print("跳过空数据行")
            continue

        # 4. 构造邮件内容
        msg = MIMEMultipart()
        msg["From"] = sender_email  # 发件人
        msg["To"] = client_email  # 收件人
        msg["Subject"] = f"【{client_name}】您的月度销售报表已发送"  # 邮件主题

        # 邮件正文（支持HTML格式，可自定义内容）
        body = f"""
        <p>您好，{client_name}！</p>
        <p>这是您本月的销售报表，已附在邮件中，请查收。</p>
        <p>如有疑问，请随时联系我~</p>
        """
        msg.attach(MIMEText(body, "html", "utf-8"))

        pdf_entire_path = os.path.join(root_path, pdf_path)
        # 5. 添加PDF附件
        try:
            with open(pdf_entire_path, "rb") as f:
                pdf_attach = MIMEApplication(f.read(), _subtype="pdf")
                # 设置附件名（显示给收件人的名称）
                pdf_attach.add_header("Content-Disposition", "attachment", filename=pdf_path.split("\\")[-1])
                msg.attach(pdf_attach)
        except Exception as e:
            print(f"添加附件失败（{client_name}）：{str(e)}")
            continue

        # 6. 发送邮件
        try:
            server.sendmail(sender_email, client_email, msg.as_string())
            print(f"邮件发送成功！收件人：{client_email}")
        except Exception as e:
            print(f"邮件发送失败（{client_email}）：{str(e)}")

    # 7. 关闭邮件服务器
    server.quit()
    print("所有邮件处理完成！")


print(f'__name__:{__name__}')
# ------------------- 调用函数（关键参数要改！） -------------------
# pyinstaller --onefile app.py
if __name__ == "__main__":
    # 1. 邮件服务器配置（以QQ邮箱为例，其他邮箱参考下表）
    SMTP_SERVER = "smtp.qq.com"  # QQ邮箱SMTP服务器
    SMTP_PORT = 465  # QQ邮箱SSL端口
    SENDER_EMAIL = "xxxxxx@qq.com"  # 你的发件邮箱
    # 注意：这里用“邮箱授权码”，不是QQ密码！获取方式：QQ邮箱→设置→账户→开启POP3/SMTP→生成授权码
    SENDER_PASSWORD = "xxxxxx"

    # 2. 客户信息Excel路径（格式：客户姓名、客户邮箱、PDF附件路径）
    root_path = input("请输入客户信息根路径：")  # "客户邮箱列表.xlsx"
    excel_path = input("请输入客户信息Excel路径：")  # "客户邮箱列表.xlsx"
    # entire_path = os.path.join(root_path, excel_path)
    # print(entire_path)
    # 3. 发送邮件
    send_auto_email(SMTP_SERVER, SMTP_PORT, SENDER_EMAIL, SENDER_PASSWORD, root_path, excel_path)
    input("任意按键，程序结束!!!!")
