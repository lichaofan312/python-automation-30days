# 从现有Excel文件中读取数据并标红指定关键词
"""
因为xlsxwriter库只能创建新的Excel文件，无法修改或读取已有的Excel文件。
因此，在方法1和方法2中我们是通过创建新的Excel文件并将数据进行关键词标红插入该文件中。
现在我们需要结合openpyxl库读取已有的Excel文件内容，
然后通过xlsxwriter库对特定内容进行标红并保存为新文件
"""
import xlsxwriter
import re
from openpyxl import load_workbook

"""
从下面可以看到，我在原始文件.xlsx中插入了一些文本内容，
在后续我会通过python处理下面文件，
将每个单元格内容中的一些关键词进行标红，
然后生成一个新的文件 标红后文件.xlsx
"""

# 读取现有Excel文件
input_file = "原始文件.xlsx"
output_file = "标红后文件.xlsx"
# 需要标红的关键词
keywords = ["重要", "紧急", "注意"]

# 通过openpyxl模块加载现有的工作簿
wb_read = load_workbook(input_file)
# 激活活动表
sheet_read = wb_read.active

# 通过xlsxwriter创建新的工作簿用于写入标红后的内容
wb_write = xlsxwriter.Workbook(output_file)
# 生成表对象
sheet_write = wb_write.add_worksheet()

# 创建红色字体格式
red_format = wb_write.add_format({'color': 'red'})

# 编译正则表达式模式，用于匹配关键词
pattern = re.compile('|'.join(keywords))

# 遍历原工作表中的单元格
for row_idx, row in enumerate(sheet_read.values):
    for col_idx, cell_value in enumerate(row):
        if cell_value is None:
            continue

        # 将单元格值转换为字符串
        content = str(cell_value)

        # 查找所有匹配的关键词
        words = re.findall(pattern, content)

        if not words:
            # 如果没有找到关键词，直接写入原文本
            sheet_write.write(row_idx, col_idx, content)
        else:
            # 如果找到关键词，生成富文本
            content_sub = re.sub(pattern, '*', content)
            formatted_content = []
            start = 0
            index = 0

            # 构建富文本参数列表
            while content_sub.find('*', start) != -1:
                end = content_sub.find('*', start)
                if start != end:
                    # 关键词前面的内容
                    formatted_content.append(content_sub[start:end])
                # 对关键词进行标红
                formatted_content.append(red_format)
                formatted_content.append(words[index])
                start = end + 1
                index += 1

            # 添加最后一部分文本
            if start < len(content_sub):
                formatted_content.append(content_sub[start:])

            # 写入富文本
            sheet_write.write_rich_string(row_idx, col_idx, *formatted_content)

# 保存并关闭工作簿
wb_write.close()
