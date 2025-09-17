# #  生成一个文件对象
# workbook = xl.Workbook()
# #  生成一个表对象
# sheet = workbook.active
#
# sheet.title = 'sheet表1'
# workbook.save('1.xlsx')

print("*" * 20)

# # 修改单元格数据，Excel表格另存为
# #  1 打开文件， 生成文件对象
# workbook = xl.load_workbook('test.xlsx')
# #  2 创建活动表对象
# sheet = workbook.active
# #  3 修改表中单元格数据
# sheet['A1'] = 'name'
# sheet['B1'].value = 'hobby'
#
# # 这里文件名使用原名，文件直接保存，
# # 若使用新的名称，文件会另存为新的文件
# workbook.save('2.xlsx')
print("*" * 20)

# # 向表中添加数据
# workbook_name = 'test.xlsx'
# workbook = xl.load_workbook(workbook_name)
# sheet = workbook.active
# # 向表中按行插入数据
# sheet.append(['李超凡', '篮球', '哈哈哈'])
# workbook.save(workbook_name)
print("*" * 20)

# # 向表中插入空白行和空白列
# workbook_name = 'test.xlsx'
# workbook = xl.load_workbook(workbook_name)
# sheet = workbook.active
#
# # 表示在第2行向下插入2个空白行
# sheet.insert_rows(idx=2, amount=2)
# sheet.insert_cols(idx=2, amount=1)
# workbook.save(workbook_name)  # 执行程序的时候， 不可以打开
print("*" * 20)

# # 删除表中的行和列
# workbook_name = 'test.xlsx'
# workbook = xl.load_workbook(workbook_name)
# sheet = workbook.active
# sheet.delete_rows(idx=2, amount=2)
# sheet.delete_cols(idx=2, amount=1)
# workbook.save(workbook_name)
print("*" * 20)

# 字母列号与数字列号之间的转换（了解）
# from openpyxl.utils import get_column_letter,column_index_from_string
# workbook_name = 'test.xlsx'
# workbook = xl.load_workbook(workbook_name)
# sheet = workbook.active
# # 3、根据列的数字返回字母
# print(get_column_letter(3))  # 结果：C ,这里相当于第3列，用字母表示就是C
# # 4、更据字母返回列数
# print(column_index_from_string('D')) #  # 结果：4 ，这里相当于D列，用数字表示就是第4列
print("*" * 20)

# 7 设置字体样式
# workbook_name = 'test.xlsx'
# workbook = xl.load_workbook(workbook_name)
# sheet = workbook.active
# # 获取单元格对象
# cell = sheet['A1']
# # 获取字体样式对象
# font = cell.font
# print('下面是该单元格数据的字体样式：')
# print(font.name, font.size, font.bold, font.italic, font.color)
# # 修改字体样式
# import openpyxl.styles as style
#
# # cell.font 一定要使用cell.font 来赋值
# cell.font = style.Font(name='微软雅黑', size='20', bold=True, italic=False, color='FF0000')
# #  保存文件
# workbook.save(workbook_name)

# 8 批量 设置字体样式

# workbook_name = 'test.xlsx'
# workbook = xl.load_workbook(workbook_name)
# sheet = workbook.active
#
# # 获取单元格对象
# cells = sheet['B']
# print(cells)
# for cell in cells:
#     cell.font = style.Font(name='微软雅黑', size='20', bold=True, italic=True, color='FF00FF')
# #  保存文件
# workbook.save(workbook_name)

# workbook_name = 'test.xlsx'
# workbook = xl.load_workbook(workbook_name)
# # 创建活动表对象
# sheet = workbook.active

# # 修改一个单元格对齐格式
# cell = sheet['A2']
# # 修改对齐格式
# cell.alignment = style.Alignment(horizontal='center', vertical='center', textRotation=0, wrap_text=True)
# workbook.save(workbook_name)

# # 修改多个单元格对齐格式
# cells = sheet["A"]
# print(cells)
# for cell in cells:
#     cell.alignment = style.Alignment(horizontal='center', vertical='center', textRotation=0, wrap_text=True)
# workbook.save(workbook_name)

# # 设置行高、列宽
# # 设置第1行 高度为60磅
# sheet.row_dimensions[1].height = 60  # 磅
# # 设置第B列的宽度为30字符
# sheet.column_dimensions['B'].width = 30  # 字符
# workbook.save(workbook_name)

# #  合并单元格
# sheet.merge_cells('B1:C1')  # 合并B1 C1 两个单元格
# sheet.merge_cells(start_row=4, end_row=5, start_column=1, end_column=2)
# workbook.save(workbook_name)

# # 拆分单元格
# sheet.unmerge_cells('B1:C1')
# workbook.save(workbook_name)

# # 设置单元格边框样式（颜色和线条）
# from openpyxl import Workbook
# from openpyxl.styles import Border, Side
#
# wb = Workbook()
# # 激活工作表
# sheet = wb.active
# # 创建一个border对象， 设置单元格上下左右各边框的样式
# border = Border(
#     left=Side(style='thin', color='FF0000'),
#     right=Side(style='medium', color='00FF00'),
#     top=Side(style='dashed', color='0000FF'),
#     bottom=Side(style='dotted', color='FFFF00')
# )
# # 将边框应用到目标单元格
# sheet['B3'].border = border
# # 写入数据到单元格， 以便查看效果
# sheet['B3'] = '阿凡'
# wb.save('cell_style.xlsx')

#  sheet表的创建、修改、复制、删除
#  打开文件
import openpyxl as xl

workbook = xl.load_workbook('test.xlsx')
#  创建
# workbook.create_sheet('66')
# print(workbook.sheetnames)
# workbook.save('test.xlsx')
# # 修改
# sheet = workbook['66']
# sheet.title = '阿凡666'
# print(workbook.sheetnames)
# workbook.save('test.xlsx')
##  复制
# sheet = workbook['阿凡666']
# workbook.copy_worksheet(sheet)
# print(workbook.sheetnames)
# workbook.save('test.xlsx')

# # 删除
# sheet = workbook['阿凡666']
# workbook.remove(sheet)
# print(workbook.sheetnames)
# workbook.save('test.xlsx')