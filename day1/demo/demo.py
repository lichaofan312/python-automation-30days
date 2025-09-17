import openpyxl as xl

#  生成Excel文件对象，查看所有sheet表 文件必须存在， 否则报错 工作簿
workbook = xl.load_workbook('test.xlsx')  # 这里文件名，也可以是文件路径，如：D:day/test.xlsx

# 2、打印test表格中所有的工作表，以列表形式返回
print(workbook.sheetnames)  # ['Sheet1', 'Sheet2', 'Sheet3']

# 获取指定表 得到表对象 sheet
sheet = workbook['Sheet1']
print(sheet)  # <Worksheet "Sheet1">

#  获取活动表对象 -----------------------------
sheet = workbook.active
print(sheet)  # <Worksheet "Sheet1"> 证明当前在使用或查看的表是 Sheet1

# 获取表格中数据所占大小 得到表格中数据占据了几行几列，如：4行2列 就表示为 A1:B4
res = sheet.dimensions
print(res)
print()

# 通过指定行和列，获得单元格对象，取得其中数据
cell1 = sheet.cell(row=2, column=1)
cell2 = sheet.cell(row=3, column=2)
print(cell1, cell1.value)  # <Cell 'Sheet1'.A2> 张三
print(cell1.row, cell1.column, cell1.coordinate)  # 2 1 A2
print(cell2, cell2.value)  # <Cell 'Sheet1'.B3> 跳

# 获得指定区间范围内的数据
cells = sheet['A1:A3']  # 得到 A1、A2、A3的单元格对象
print(f'所有单元格对象： {cells}')  # ((<Cell 'Sheet1'.A1>,), (<Cell 'Sheet1'.A2>,), (<Cell 'Sheet1'.A3>,))
# 4、打印出所有单元格对象中的数据，这里直接使用for循环所有对象遍历出来
print('A1:A3的单元格数据依次为：')
for cell in cells:  #
    print(cell)  # (<Cell 'Sheet1'.A1>,)
    for realCell in cell:
        print(realCell.value)

# 获取指定行或列的数据
# 1 得到指定行中所有单元格对象
cells = sheet['2']
print(f'所有单元格对象：{cells}')  # (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>)
for cell1 in cells:
    print(cell1.value)

# 2 获取某两列的单元格数据
cells2 = sheet['A:B']
print(f'所有单元格对象：{cells2}')

for cell in cells2:
    for realCell in cell:
        print(realCell.value)

# 按行或按列读取单元格数据（迭代器）
# 按行的顺序读取数据
data1 = sheet.iter_rows(min_row=2, max_row=4, min_col=1, max_col=2)
print(data1)  # # data接收的是这块区间单元格对象
for i in data1:
    for x in i:
        print(x.value)

# 获取表中数据所占的行列数
print(f'表中行有: {sheet.max_row}')  # 4
print(f'表中列有: {sheet.max_column}')  # 2
# 获取到表中所有行和列的数据
print(f'表中行有: {sheet.rows}')  # 4
print('----')
for rowCell in sheet.rows:
    for real_cell in rowCell:
        print(real_cell.value, sep='', end="|")
    print()
print('----')
print(f'表中列有: {sheet.columns}')  # 2
for columnCell in sheet.columns:
    for real_cell in columnCell:
        print(real_cell.value, sep='', end="|")
    print()
print('----')
