import openpyxl as xl

# # 创建excel对象
# wb = xl.Workbook(write_only=True) # 设置write_only=True模式，忽略数据结构
#
# # 创建sheet表，生成表对象
# ws = wb.create_sheet()
#
# # 创建一个生成器来生成数据
# data = (
#     ('唱', '跳', 'rap'),
#     (6, 66, 666),
#     ('哈', '哈哈哈', 'hhhahah')
# )
#
# # 将数据依次遍历存入excel表格中
# for row in data:
#     ws.append(row)
#
# wb.save("write_data.xlsx")

# 加快数据读取速度（read_only模式和迭代器的使用）
wb = xl.load_workbook('write_data.xlsx', read_only=True)
sheet = wb.active
# 使用迭代器按行读取数据
for row in sheet.iter_rows(min_row=1, max_row=3, min_col=1, max_col=3):
    for cell in row:
        print(cell.value)

# 3 禁用公式计算
# 设置data_only=True，禁用公式计算
# 设置keep_vba=False，关闭格式和样式
wb = xl.load_workbook('write_data.xlsx', read_only=True, data_only=True, keep_vba=True)

